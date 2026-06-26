#!/usr/bin/env python3
"""
全量重算 KPI: 基于 CSV 格式的 AData 快照
支持的 CSV 文件:
  - mapping_3xACIv_response.csv: OPS主体名 / 客户简称 / 消耗
  - full_quarter_33VMRV_response.csv: 客户简称 / 季度(2025/Q1格式) / 消耗
  - daily_E0fBs_response.csv: 客户简称 / 日均消耗 / 对比值 / 变化量 / 变化率
  - delivery_adq_15yT2p_response.csv: 客户简称 / ADQ消耗 / 环比
  - delivery_qyt_1KTroA_response.csv: 客户简称 / 全域通消耗 / 环比
  - topkey_4yCkqO_response.csv: 客户简称 / 竞价消耗 / 消耗环比 / ROI / ...
  - topstatus_20S1yv_response.csv: 客户简称 / 消耗 / ...
输出: center_quarter_summary.js, top80_effective_metrics.js, delivery_side_summary.js
用法: python3 scripts/rebuild_kpi_csv.py
"""
import csv, json, re, os, glob, sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime

PROJECT = "/Users/duziqing/WorkBuddy/2026-05-12-task-5"
# 2026-05-27 切换到 GitHub Pages 真源（单一真源死规矩，禁止再写废弃镜像）
PC_DATA = os.environ.get("REBUILD_OUTPUT_DIR", f"{PROJECT}/github_pages_adq_publish/sales-center/data")
T_MINUS_1 = "2026-06-25"

# ============================================================
# 新客季度日均锚定 AData 整体聚合（2026-05-20 用户拍板）
# 每日把 AData 截图/接口的「整体 Q2 消耗」填到这里，看板合计行 quarterCost
# 直接用这个值，避免 full_quarter 客户级聚合因写入延迟而偏低。
# 来源示例：AData 整体接口截图，截止 T-1 的 2026Q2 累计消耗。
# 留 0 时退回 full_quarter 客户级 SUM。
# ============================================================
Q2_ADATA_TOTAL_OVERRIDE = 0  # 2026-05-23 用户拍板 B：注释掉 override，让脚本从 full_quarter_response.csv 自己加总（旧值 33368589.21 截止 5/19 已过期 4 天）
# 自动算前一天 T-2，用于读快照算累计日Δ
from datetime import datetime as _dt, timedelta as _td
_t1 = _dt.strptime(T_MINUS_1, "%Y-%m-%d")
T_MINUS_2 = (_t1 - _td(days=1)).strftime("%Y-%m-%d")

def latest_adata_dir():
    # 2026-06-13 修复：优先读项目目录的 adata_refresh_*（真源），
    # /tmp/kanban_0603 仅作历史兼容兜底。之前硬编码只读 /tmp 导致
    # 项目里归档的当日 CSV 全程没被使用，看板吃旧数据（服饰中心日耗显示旧值）。
    proj = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dirs = sorted(glob.glob(os.path.join(proj, "adata_refresh_2*")))
    if dirs:
        return dirs[-1]
    dirs = sorted(glob.glob("/tmp/kanban_0603/adata_refresh_*"))
    if not dirs:
        sys.exit("No adata_refresh_* dir found")
    return dirs[-1]

def read_csv_utf8(path):
    """读取UTF-8-sig CSV，返回 list of dict"""
    rows = []
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    return rows

def clean_str(v):
    if v is None: return ""
    return str(v).replace('\n', ' ').replace('\r', ' ').replace('\t', ' ').strip()

def safe_float(v, default=0.0):
    try:
        return float(str(v).replace(',', '').strip())
    except:
        return default

def clean_rate(v):
    """2026-06-02 源头清洗：adata「环比变化率」对小基数客户(前天消耗几百元)会爆成
    -37.75、+154 这类天文数字(跌涨>100%物理不可能)。|v|>1 即脏值，置 None。
    入参 v 已是小数(原始%÷100)。"""
    if v is None:
        return None
    return None if abs(v) > 1 else v

def round_or_none(v, n):
    """None 安全 round：脏值清成 None 后下游 round 不崩。"""
    return None if v is None else round(v, n)

def clean_pct(v):
    """2026-06-02 百分数口径环比率清洗(未÷100,如 -37.75 已是 %)：
    |v|>100 即跌涨>100% 物理不可能，置 None。用于小红点 top_status_data。"""
    if v is None:
        return None
    return None if abs(v) > 100 else v

def to_quarter_from_str(s):
    """2025/Q1 -> 2025Q1 / 2025-01-01 -> 2025Q1"""
    s = str(s).strip()
    m = re.match(r'(\d{4})/Q(\d)', s)
    if m:
        return f"{m.group(1)}Q{m.group(2)}"
    m2 = re.match(r'(\d{4})-(\d{2})', s)
    if m2:
        y, mo = int(m2.group(1)), int(m2.group(2))
        q = (mo - 1) // 3 + 1
        return f"{y}Q{q}"
    return ""

print("=" * 60)
print("销售作战中心 KPI 重建 (CSV 口径)")
print(f"T-1 = {T_MINUS_1}")
print("=" * 60)

adata_dir = Path(latest_adata_dir())
print(f"AData 目录: {adata_dir}")

# ============================================================
# 1. 24 年老客
# ============================================================
OLD24_SET = set()
old24_csvs = sorted(glob.glob(f"{PROJECT}/data_fixed/24\u5e74\u8001\u5ba2-*.csv"))
if not old24_csvs:
    sys.exit("No 24年老客 CSV found")
old24_csv = old24_csvs[-1]
with open(old24_csv, encoding='utf-8') as f:
    reader = csv.reader(f)
    next(reader)
    for row in reader:
        # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
        if row[0].strip() in ('0', '1', '2'):
            row = row[1:]
        if row and row[0].strip():
            OLD24_SET.add(row[0].strip())
print(f"1. 24年老客: {len(OLD24_SET)}")

# ============================================================
# 2. 新锐名单（2026-05-27 用户拍板：新锐 = 有效新客 且 客户简称 ∈ 新锐名单）
#     2026-06-11 子青死规矩升级「全部都是每天更新」：
#     优先级 1 = ~/Downloads/ 最新 `新锐标签客户-*.xlsx`（按 mtime 取最新）
#     优先级 2 = adata_dir 内 `新锐名单_20260527.txt`（兼容旧路径）
#     优先级 3 = 标签命中 fallback（已弃用）
# ============================================================
RISING_SET = set()
loaded_from = None

# 优先级 1：~/Downloads 最新新锐名单/新锐标签客户 xlsx（每天会有新版）
import openpyxl as _openpyxl
_dl_dir = os.path.expanduser('~/Downloads')
def _is_rising_xlsx(fn):
    # 2026-06-23 修复：排除「潜力新锐名单.xlsx」这类待申请/潜力池文件，不能当已认证新锐名单。
    if not fn.endswith('.xlsx') or '潜力' in fn:
        return False
    return (
        fn.startswith('新锐标签客户-')
        or fn.startswith('新锐名单-')
    )
_rising_xlsx_candidates = sorted(
    [os.path.join(_dl_dir, fn) for fn in os.listdir(_dl_dir) if _is_rising_xlsx(fn)],
    key=lambda p: os.path.getmtime(p),
    reverse=True,
)
if _rising_xlsx_candidates:
    _xlsx = _rising_xlsx_candidates[0]
    _wb = _openpyxl.load_workbook(_xlsx, read_only=True, data_only=True)
    _ws = _wb.active
    for _row in _ws.iter_rows(values_only=True):
        if not _row: continue
        _v = _row[0]
        if not _v: continue
        _name = str(_v).strip()
        if _name and _name != '客户简称':
            RISING_SET.add(_name)
    _wb.close()
    loaded_from = os.path.basename(_xlsx)
    print(f"2. 新锐名单: {len(RISING_SET)}  (source=~/Downloads/{loaded_from})")

# 优先级 2：adata_dir 内 txt（兼容旧）
if not RISING_SET:
    rising_txt = os.path.join(str(adata_dir), '新锐名单_20260527.txt')
    if os.path.exists(rising_txt):
        with open(rising_txt, 'r', encoding='utf-8') as f:
            for line in f:
                name = line.strip()
                if name:
                    RISING_SET.add(name)
        loaded_from = '新锐名单_20260527.txt'
        print(f"2. 新锐名单: {len(RISING_SET)}  (source={loaded_from})")

# 优先级 3：标签命中 fallback
if not RISING_SET:
    rising_path = '/Users/duziqing/WorkBuddy/20260506112439/output/'
    RISING_TAGS = {'新锐', '经销商'}
    candidate = []
    if os.path.exists(rising_path):
        for fn in sorted(os.listdir(rising_path), reverse=True):
            if not fn.endswith('.xlsx'): continue
            if '客户标签' in fn or '客群标签' in fn or '新锐' in fn:
                candidate.append(fn)
    for fn in candidate:
        full = os.path.join(rising_path, fn)
        wb = _openpyxl.load_workbook(full, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None) or ()
        tag_idx = 1 if headers and len(headers) <= 2 else 3
        cnt_before = len(RISING_SET)
        for row in rows_iter:
            if not row: continue
            short = str(row[0] or '').strip()
            if not short: continue
            tag_raw = str(row[tag_idx] or '') if tag_idx < len(row) else ''
            if any(t in tag_raw for t in RISING_TAGS):
                RISING_SET.add(short)
        wb.close()
        print(f"   load {fn} (tag_idx={tag_idx}) +{len(RISING_SET)-cnt_before}")
        loaded_from = fn
        break
    if RISING_SET:
        print(f"2. 新锐标签(fallback): {len(RISING_SET)}  (source={loaded_from})")


# ============================================================
# 3. mapping: 主体 -> 简称
# ============================================================
SUBJECT_SHORT_MAP = {}  # {ops_name: short}
mapping_file = adata_dir / "mapping_3xACIv_response.csv"
if mapping_file.exists():
    rows = read_csv_utf8(str(mapping_file))
    for row in rows:
        keys = list(row.keys())
        if len(keys) < 2: continue
        ops = clean_str(row[keys[0]])
        short = clean_str(row[keys[1]])
        if ops and short and ops != '\u6574\u4f53':
            SUBJECT_SHORT_MAP[ops] = short
# 2026-06-22 修复：新锐名单里有些是旧客户简称，最新 mapping 已把同一主体映射到新简称。
# 例如「杭州卡路里体育有限公司」→「Keep」。若只按旧简称匹配，AData full_quarter 用新简称时会漏算新锐。
_rising_alias_added = []
for _ops, _short in SUBJECT_SHORT_MAP.items():
    if _ops in RISING_SET and _short and _short not in RISING_SET:
        RISING_SET.add(_short)
        _rising_alias_added.append((_ops, _short))
print(f"3. 主体->简称映射: {len(SUBJECT_SHORT_MAP)}")
if _rising_alias_added:
    print(f"   新锐简称别名补齐 {len(_rising_alias_added)} 个：" + ", ".join([f"{a}->{b}" for a,b in _rising_alias_added[:20]]))

# ============================================================
# 4. full_quarter: 简称 -> {季度: 消耗}
# ============================================================
SHORT_QUARTER = defaultdict(dict)  # {short: {q: cost}}
fq_file = adata_dir / "full_quarter_33VMRV_response.csv"
if not fq_file.exists():
    sys.exit("full_quarter_33VMRV_response.csv not found")
with open(str(fq_file), encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    next(reader)  # header
    for row in reader:
        # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
        if row[0].strip() in ('0', '1', '2'):
            row = row[1:]
        if len(row) < 3: continue
        # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
        if row[0].strip() in ('0', '1', '2'):
            short = row[1].strip()
            q_raw = row[2].strip()
            cost_str = row[3].strip() if len(row) > 3 else ''
        else:
            short = row[0].strip()
            q_raw = row[1].strip()
            cost_str = row[2].strip()
        if not short or short == '客户简称' or not q_raw: continue
        cost = safe_float(cost_str)
        if cost <= 0: continue
        q_norm = to_quarter_from_str(q_raw)
        if q_norm:
            SHORT_QUARTER[short][q_norm] = SHORT_QUARTER[short].get(q_norm, 0) + cost
print(f"4. 全量季度消耗 客户数: {len(SHORT_QUARTER)}")

# ============================================================
# 5. daily: 简称 -> {yest, prev}
# ============================================================
DAILY = {}
daily_file = adata_dir / "daily_E0fBs_response.csv"
if not daily_file.exists():
    sys.exit("daily_E0fBs_response.csv not found")
with open(str(daily_file), encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    next(reader)
    for row in reader:
        # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
        if row[0].strip() in ('0', '1', '2'):
            row = row[1:]
        if len(row) < 3: continue
        short = row[0].strip()
        if not short or short == '\u6574\u4f53': continue
        yest = safe_float(row[1])
        prev = safe_float(row[2]) if len(row) > 2 else 0
        # 直接用 CSV 自带的环比变化率（第 4 列），不自己算
        rate = safe_float(row[4]) / 100 if len(row) > 4 else 0
        DAILY[short] = {'yest': yest, 'prev': prev, 'rate': rate}
DAILY_TOTAL_YEST = 0
DAILY_TOTAL_PREV = 0
DAILY_TOTAL_RATE = 0
with open(str(daily_file), encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    next(reader)
    for row in reader:
        # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
        if row[0].strip() in ('0', '1', '2'):
            row = row[1:]
        if len(row) >= 2 and (row[0].strip() == '\u6574\u4f53' or row[0].strip() == ''):
            DAILY_TOTAL_YEST = safe_float(row[1])
            DAILY_TOTAL_PREV = safe_float(row[2]) if len(row) > 2 else 0
            DAILY_TOTAL_RATE = safe_float(row[4]) / 100 if len(row) > 4 else 0
            break
print(f"5. daily\u5ba2\u6237\u6570: {len(DAILY)}  \u670d\u9970\u4e2d\u5fc3\u6628\u65e5\u6d88\u8017: {DAILY_TOTAL_YEST/10000:.1f}\u4e07  \u524d\u65e5: {DAILY_TOTAL_PREV/10000:.1f}\u4e07  CSV\u73af\u6bd4: {DAILY_TOTAL_RATE*100:+.2f}%")

# ============================================================
# 6. 投放端 ADQ / 全域通
# ============================================================
ADQ_SET = set()
adq_file = adata_dir / "delivery_adq_15yT2p_response.csv"
if adq_file.exists():
    with open(str(adq_file), encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
            if row[0].strip() in ('0', '1', '2'):
                row = row[1:]
            s = row[0].strip() if row else ''
            if s and s != '\u6574\u4f53': ADQ_SET.add(s)
print(f"6. ADQ投放端客户数: {len(ADQ_SET)}")

QYT_SET = set()
qyt_file = adata_dir / "delivery_qyt_1KTroA_response.csv"
if qyt_file.exists():
    with open(str(qyt_file), encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
            if row[0].strip() in ('0', '1', '2'):
                row = row[1:]
            s = row[0].strip() if row else ''
            if s and s != '\u6574\u4f53': QYT_SET.add(s)
print(f"6. 全域通客户数: {len(QYT_SET)}")

DELIVERY_SIDE_MAP = {}
for s in (ADQ_SET | QYT_SET):
    if s in ADQ_SET and s in QYT_SET:
        DELIVERY_SIDE_MAP[s] = '\u53cc\u6295\u653e'
    elif s in ADQ_SET:
        DELIVERY_SIDE_MAP[s] = 'ADQ'
    else:
        DELIVERY_SIDE_MAP[s] = '\u5168\u57df\u901a'

# ============================================================
# ===========================================================
# 7. records (tuoke_real_records.js) - 读取登记底表
# ===========================================================
# 优先读 /tmp/kanban_0603/clean_tuoke_base.json（干净双层底表：12733全量基线 + 471登记层带_createdAt）
# ⚠️ 旧的 /tmp/merged_tuoke_records.json 已弃用（脏：id/_id字段不统一无法去重，时间戳覆盖仅6%）
merged_json = '/tmp/kanban_0603/clean_tuoke_base.json'
if os.path.exists(merged_json):
    with open(merged_json, encoding='utf-8') as f:
        records = json.load(f)
    print(f"7. 拓新登记底表 records: {len(records)} (source=clean_tuoke_base.json 干净双层底表)")
else:
    records_js = f"{PC_DATA}/tuoke_real_records.js"
    if os.path.exists(records_js):
        with open(records_js) as f:
            s = f.read()
        m = re.search(r'__TUOKE_REAL_RECORDS__\s*=\s*(\[[\s\S]*?\]);', s)
        if m:
            records = json.loads(m.group(1))
    print(f"7. 拓新登记底表 records: {len(records)} (source=tuoke_real_records.js)")
SALES = ["brownfan", "Jonzhu", "kaikaigenli", "kinsleyjin", "lijunwu", "ruilingzhan", "yvaineechen"]

# ===========================================================
# 人工归属覆盖层（manual_attr_override.json）：
#   子青指定销售归属最高优先级。已有客户原地改写 sale/_rtx/_recorded_by，
#   仅登记底表没有的客户才合成 _manual_override 记录，避免全量归属表导致底表膨胀。
# ===========================================================
# 2026-06-05：override 文件迁到项目 data 目录持久化（纳入git，永不丢失）。
# 优先读项目真源，兜底 /tmp（兼容旧路径）。
_override_path = f'{PROJECT}/github_pages_adq_publish/sales-center/data/manual_attr_override.json'
if not os.path.exists(_override_path):
    _override_path = '/tmp/kanban_0603/manual_attr_override.json'
if os.path.exists(_override_path):
    with open(_override_path, encoding='utf-8') as f:
        _ov = json.load(f).get('overrides', {})
    if _ov:
        import time as _t
        # 2026-06-16 修复：每次 rebuild 前先移除上一轮生成的 _manual_override 记录，
        # 避免同一批人工归属覆盖在 tuoke_real_records.js 里越跑越多。
        records = [r for r in records if not r.get('_manual_override')]
        _now_ms = int(_t.time() * 1000)
        _by_short_override = {}
        for _r in records:
            _short_key = (_r.get('shortName') or _r.get('name') or '').strip()
            if _short_key:
                _by_short_override.setdefault(_short_key, []).append(_r)
        _applied = []
        _synthetic = 0
        for _short, _info in _ov.items():
            _target = (_info.get('sale') or '').strip()
            if not _target:
                continue
            _hits = _by_short_override.get(_short, [])
            if _hits:
                # 2026-06-18 子青给 1.1w 全量归属表：已有记录直接原地改 sale，
                # 不再为每条覆盖追加副本，避免 tuoke_real_records.js 成倍膨胀。
                for _r in _hits:
                    _r['sale'] = _target
                    _r['_rtx'] = _target
                    _r['_recorded_by'] = _target
                    _r['saleEffectiveAt'] = _now_ms
                    _r['_updatedAt'] = _now_ms
                    _r['_manual_override_source'] = 'manual_attr_override'
            else:
                # AData 里有但登记底表还没有的客户，追加一条极简合成记录，
                # 让 KPI/Top80 也能按用户指定销售归属统计。
                _tmpl = {'shortName': _short, 'name': _short}
                _tmpl['sale'] = _target
                _tmpl['_rtx'] = _target
                _tmpl['_recorded_by'] = _target
                _tmpl['_createdAt'] = _now_ms
                _tmpl['_updatedAt'] = _now_ms
                _tmpl['saleEffectiveAt'] = _now_ms
                _tmpl['_manual_override'] = True
                records.append(_tmpl)
                _synthetic += 1
            _applied.append((_short, _target))
        if _applied:
            print(f"   人工归属覆盖 {len(_applied)} 个（合成{_synthetic}个，其余原地改写）")
            for _s, _t2 in _applied[:30]:
                print(f"     {_s} -> {_t2}")
            if len(_applied) > 30:
                print(f"     ... 省略 {len(_applied)-30} 个")

# ===========================================================
# 归属映射（双层模型）：
#   TUOKE_SHORT_TO_SALE_FIRST = 先登记先归属（按真实写入时间取最早，永久存证、防抢客）
#   TUOKE_SHORT_TO_SALE       = 当前生效归属（按真实写入时间取最新，反映合法交接/更替）
# ⚠️ 排序必须用 _createdAt（真实写入时间戳），绝不能用 date 字段——
#    改归属时前端会把 date 写成客户首投季度对应的过去日期，比原记录更早，按 date 排序会把"最新修改"误判成"最旧"。
# ===========================================================
def _ctime(r):
    # 2026-06-06 归属生效时间戳优先：前端改归属时写 saleEffectiveAt=Date.now()。
    # 因 cloud.upload 对同 id 是 update 覆盖、_createdAt 不变，双层模型若仅按 _createdAt
    # 排序会认不出"最新改的归属"。故归属排序键取 max(saleEffectiveAt, _createdAt, _updatedAt)，
    # 让带 saleEffectiveAt 的改动一定排到最新 → 当前生效归属取到改后的销售。
    cand = []
    for k in ("saleEffectiveAt", "_createdAt", "_updatedAt"):
        v = r.get(k)
        if isinstance(v, (int, float)):
            cand.append(v)
    return max(cand) if cand else 0

_by_short = {}
for r in records:
    short = (r.get("shortName") or r.get("name") or "").strip()
    sale = (r.get("sale") or "").strip()
    if short and sale:
        _by_short.setdefault(short, []).append(r)

TUOKE_SHORT_TO_SALE = {}        # 当前生效归属（最新）
TUOKE_SHORT_TO_SALE_FIRST = {}  # 先登记归属（最早）
_changed_attr = []              # 发生过归属更替的客户（最早≠最新）
for short, rs in _by_short.items():
    rs_sorted = sorted(rs, key=_ctime)
    first_sale = (rs_sorted[0].get("sale") or "").strip()
    last_sale = (rs_sorted[-1].get("sale") or "").strip()
    TUOKE_SHORT_TO_SALE_FIRST[short] = first_sale
    TUOKE_SHORT_TO_SALE[short] = last_sale
    if first_sale != last_sale:
        _changed_attr.append((short, first_sale, last_sale))
if _changed_attr:
    print(f"   归属更替客户 {len(_changed_attr)} 个（先登记→现生效）：")
    for short, fs, ls in _changed_attr:
        print(f"     {short}: {fs} → {ls}")

# ===========================================================
# 8. 核心 KPI 计算
# ============================================================
ALL_SHORTS = set(SHORT_QUARTER.keys())
print(f"\n服饰中心全量客户: {len(ALL_SHORTS)}")

def compute_first_quarter(short):
    qs = SHORT_QUARTER.get(short, {})
    if not qs: return None
    return min(qs.keys())

def quarter_cost(short, q):
    return SHORT_QUARTER.get(short, {}).get(q, 0)

def compute_segment(target_q, sale_filter=None):
    new_count = valid_count = rising_count = 0
    yest_cost = prev_cost = 0.0
    quarter_cost_sum = 0.0  # 季度累计消耗（用于算日均）

    for short in ALL_SHORTS:
        if short in OLD24_SET: continue
        fq = compute_first_quarter(short)
        if fq != target_q: continue
        # 销售筛选
        owner = TUOKE_SHORT_TO_SALE.get(short)
        if sale_filter is None:
            pass
        elif sale_filter == 'others':
            if owner in SALES: continue
        else:
            if owner != sale_filter: continue

        new_count += 1
        qc = quarter_cost(short, target_q)
        quarter_cost_sum += qc
        is_valid = qc > 1000
        if is_valid:
            valid_count += 1
        if is_valid and short in RISING_SET:
            rising_count += 1
        d = DAILY.get(short, {})
        yest_cost += d.get('yest', 0)
        prev_cost += d.get('prev', 0)

    day_rate = (yest_cost - prev_cost) / prev_cost if prev_cost > 0 else 0
    return {
        'newCount': new_count,
        'validCount': valid_count,
        'risingCount': rising_count,
        'yestCost': yest_cost,
        'prevCost': prev_cost,
        'dayCostRate': day_rate,
        'quarterCost': quarter_cost_sum
    }

# ============================================================
# 9. 生成 center_quarter_summary.js
# ============================================================
target_quarters = ["2026Q2", "2026Q1", "2025Q4", "2025Q3", "2025Q2", "2025Q1"]
prev_q_map = {
    "2026Q2": "2026Q1", "2026Q1": "2025Q4", "2025Q4": "2025Q3",
    "2025Q3": "2025Q2", "2025Q2": "2025Q1", "2025Q1": None
}

new_summary = []
print("\n=== 2026Q2 KPI 明细 ===")

for tq in target_quarters:
    pq = prev_q_map[tq]
    each_sales = {sale: compute_segment(tq, sale) for sale in SALES}
    other = compute_segment(tq, 'others')
    total = compute_segment(tq, None)

    if pq:
        prev_each = {sale: compute_segment(pq, sale) for sale in SALES}
        prev_other = compute_segment(pq, 'others')
        prev_total = compute_segment(pq, None)
    else:
        z = {'newCount': 0, 'validCount': 0, 'risingCount': 0, 'yestCost': 0, 'prevCost': 0, 'dayCostRate': 0}
        prev_each = {sale: dict(z) for sale in SALES}
        prev_other = dict(z)
        prev_total = dict(z)

    if tq == "2026Q2":
        seven_sum = {
            'newCount': sum(v['newCount'] for v in each_sales.values()),
            'validCount': sum(v['validCount'] for v in each_sales.values()),
            'risingCount': sum(v['risingCount'] for v in each_sales.values()),
            'yestCost': sum(v['yestCost'] for v in each_sales.values()),
        }
        print(f"  增长组(7销售): 新客 {seven_sum['newCount']} / 有效 {seven_sum['validCount']} / 新锐 {seven_sum['risingCount']} / 昨耗 {seven_sum['yestCost']/10000:.1f}万")
        print(f"  其他档:        新客 {other['newCount']} / 有效 {other['validCount']} / 新锐 {other['risingCount']} / 昨耗 {other['yestCost']/10000:.1f}万")
        print(f"  服饰中心合计:  新客 {total['newCount']} / 有效 {total['validCount']} / 新锐 {total['risingCount']} / 昨耗 {total['yestCost']/10000:.1f}万 / 日环比 {total['dayCostRate']*100:+.1f}%")
        # 季度日均消耗对账（AData 整体覆盖 vs full_quarter 客户级）
        from datetime import date as _d
        _q2_start = _d(2026, 4, 1)
        _t1_d = _d(*[int(x) for x in T_MINUS_1.split('-')])
        _q2_days = (_t1_d - _q2_start).days  # 截止 T-1 - 1 = T-2，已过天数 = T-1 - 4/1
        _qcost_used = Q2_ADATA_TOTAL_OVERRIDE if Q2_ADATA_TOTAL_OVERRIDE > 0 else total['quarterCost']
        _avg_used = _qcost_used / _q2_days / 10000 if _q2_days > 0 else 0
        _avg_raw = total['quarterCost'] / _q2_days / 10000 if _q2_days > 0 else 0
        print(f"  Q2 季度日均: AData 锚定 {_qcost_used/10000:.2f}万 ÷ {_q2_days} 天 = {_avg_used:.2f}万/天")
        print(f"             (full_quarter 客户级 {total['quarterCost']/10000:.2f}万 → {_avg_raw:.2f}万/天)")
        if pq:
            print(f"  vs {pq}: 新客 {prev_total['newCount']} / 有效 {prev_total['validCount']} / 新锐 {prev_total['risingCount']}")
        for sale in SALES:
            c = each_sales[sale]
            print(f"    {sale}: 新客 {c['newCount']} / 有效 {c['validCount']} / 新锐 {c['risingCount']} / 昨耗 {c['yestCost']/10000:.1f}万")

    for sale in SALES:
        c, p = each_sales[sale], prev_each[sale]
        new_summary.append({
            "quarter": tq, "quarterLabel": tq.replace("Q", " Q"), "sale": sale,
            "newCount": c["newCount"], "newDelta": float(c["newCount"] - p["newCount"]),
            "newDayDelta": 0 if tq == "2026Q2" else None,
            "validCount": c["validCount"], "validDelta": float(c["validCount"] - p["validCount"]),
            "validDayDelta": 0 if tq == "2026Q2" else None,
            "risingCount": c["risingCount"], "risingDelta": float(c["risingCount"] - p["risingCount"]),
            "risingDayDelta": 0 if tq == "2026Q2" else None,
            "yestCost": c["yestCost"],
            "costRate": (c["yestCost"]-p["yestCost"])/p["yestCost"] if p["yestCost"] > 0 else None,
            "dayCostRate": c["dayCostRate"],
            "quarterCost": c["quarterCost"]
        })

    new_summary.append({
        "quarter": tq, "quarterLabel": tq.replace("Q", " Q"), "sale": "\u5176\u4ed6",
        "newCount": other["newCount"], "newDelta": float(other["newCount"] - prev_other["newCount"]),
        "newDayDelta": 0 if tq == "2026Q2" else None,
        "validCount": other["validCount"], "validDelta": float(other["validCount"] - prev_other["validCount"]),
        "validDayDelta": 0 if tq == "2026Q2" else None,
        "risingCount": other["risingCount"], "risingDelta": float(other["risingCount"] - prev_other["risingCount"]),
        "risingDayDelta": 0 if tq == "2026Q2" else None,
        "yestCost": other["yestCost"],
        "costRate": (other["yestCost"]-prev_other["yestCost"])/prev_other["yestCost"] if prev_other["yestCost"] > 0 else None,
        "dayCostRate": other["dayCostRate"],
        "quarterCost": other["quarterCost"]
    })

    new_summary.append({
        "quarter": tq, "quarterLabel": tq.replace("Q", " Q"), "sale": "\u5408\u8ba1",
        "newCount": total["newCount"], "newDelta": float(total["newCount"] - prev_total["newCount"]),
        "newDayDelta": 0 if tq == "2026Q2" else None,
        "validCount": total["validCount"], "validDelta": float(total["validCount"] - prev_total["validCount"]),
        "validDayDelta": 0 if tq == "2026Q2" else None,
        "risingCount": total["risingCount"], "risingDelta": float(total["risingCount"] - prev_total["risingCount"]),
        "risingDayDelta": 0 if tq == "2026Q2" else None,
        "yestCost": total["yestCost"],
        "costRate": (total["yestCost"]-prev_total["yestCost"])/prev_total["yestCost"] if prev_total["yestCost"] > 0 else None,
        "dayCostRate": total["dayCostRate"],
        "quarterCost": (Q2_ADATA_TOTAL_OVERRIDE if (tq == "2026Q2" and Q2_ADATA_TOTAL_OVERRIDE > 0) else total["quarterCost"]),
        "quarterCostRaw": total["quarterCost"]
    })

# 把日Δ注入到 2026Q2 合计行（如果 T-2 快照存在）
# 2026-05-22 修复：用户每天给的 adata_refresh_<T-1>/ 里只有当天 full_quarter_response.csv，
# 不会留昨天的 full_quarter_<T-2>.csv 副本，所以原逻辑全部走不到 → 日Δ 永远 0。
# 修复策略：优先用 adata_dir 的 T-2 副本（老逻辑兼容）；兜底从 PC_DATA/snapshots/<T-2>/center_quarter_summary.js
# 直接读 2026Q2 各销售/其他/合计行的 newCount/validCount/risingCount 当基准（这是 T-2 自己 rebuild 时落地的快照）
prev_day_snap_file = adata_dir / f"full_quarter_{T_MINUS_2}.csv"
# 2026-05-25 修复：T-2 不存在时（某天没跑看板）向前回溯最多 7 天找最近一个 snapshot
prev_kpi_snap_file = Path(f"{PC_DATA}/snapshots/{T_MINUS_2}/center_quarter_summary.js")
PREV_SNAP_DATE = T_MINUS_2
if not prev_kpi_snap_file.exists():
    # 2026-06-15：snapshots 已从 Pages 移出，脚本侧本地目录仍可作为日Δ基准。
    # 2026-06-16 修复：回溯必须从 T-2(前一天) 开始，不能从 2 天前开始，否则会跳过昨天快照导致日Δ按前多天累计。
    _snap_roots = [
        Path(f"{PC_DATA}/snapshots"),
        Path(f"{PROJECT}/销售作战中心_v3.2_2026-05-12/data/snapshots"),
        Path(f"{PROJECT}/github_pages_adq_publish/sales-center-mobile/data/snapshots"),
        Path(f"{PROJECT}/adq-backend/public/sales-center/data/snapshots"),
    ]
    for _back in range(1, 8):
        _cand_date = (_t1 - _td(days=_back)).strftime("%Y-%m-%d")
        for _root in _snap_roots:
            _cand = _root / _cand_date / "center_quarter_summary.js"
            if _cand.exists():
                prev_kpi_snap_file = _cand
                PREV_SNAP_DATE = _cand_date
                print(f"  ⚠️ snapshot {T_MINUS_2}/ 不存在，回溯到 {_cand_date}/ 算日Δ（间隔 {_back} 天, source={_root})")
                break
        if prev_kpi_snap_file.exists():
            break
# 2026-06-26 子青拍板"强算环比": snapshot 优先 (snapshot 是 T-2 当天落地的真实测算, 同口径)
# T-2 full_quarter 反算用今日 RISING_SET 会失真 (kaikaigenli 今天 28 vs 反算 17 → -1 不合理)
if prev_kpi_snap_file.exists():
    # 兜底：直接从 T-2 自己的 snapshot center_quarter_summary.js 读 2026Q2 行的 newCount/validCount/risingCount
    import re as _re
    snap_text = prev_kpi_snap_file.read_text(encoding='utf-8')
    m = _re.search(r"window\.__CENTER_QUARTER_SUMMARY__\s*=\s*(\[.*?\]);", snap_text, _re.S)
    prev_summary = json.loads(m.group(1)) if m else []
    prev_by_sale = {}
    for r in prev_summary:
        if r.get("quarter") != "2026Q2": continue
        prev_by_sale[r.get("sale", "")] = r
    def _clamp_pos(v):
        return v  # 不再clamp，如实显示（含负数）
    for entry in new_summary:
        if entry.get("quarter") != "2026Q2": continue
        sale = entry.get("sale")
        prev = prev_by_sale.get(sale)
        if not prev: continue
        entry["newDayDelta"] = _clamp_pos(int(entry["newCount"]) - int(prev.get("newCount", 0)))
        entry["validDayDelta"] = _clamp_pos(int(entry["validCount"]) - int(prev.get("validCount", 0)))
        entry["risingDayDelta"] = _clamp_pos(int(entry["risingCount"]) - int(prev.get("risingCount", 0)))
    print(f"\n  \u65e5\u0394 \u5df2\u6ce8\u5165 center_quarter_summary.js (\u5408\u8ba1\u884c + 7\u9500\u552e + \u5176\u4ed6) - \u8d70 snapshot \u5151\u5e95\u5206\u652f T-2={PREV_SNAP_DATE}")
elif prev_day_snap_file.exists():
    prev_short_quarter = defaultdict(dict)
    with open(str(prev_day_snap_file), encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
            if row[0].strip() in ('0', '1', '2'):
                row = row[1:]
            if len(row) < 3: continue
            short = row[0].strip()
            if not short or short == '\u6574\u4f53': continue
            q = to_quarter_from_str(row[1])
            cost = safe_float(row[2])
            if q and cost > 0:
                prev_short_quarter[short][q] = prev_short_quarter[short].get(q, 0) + cost
    pn_total = pv_total = pr_total = 0
    pn_each = {sale: 0 for sale in SALES}
    pv_each = {sale: 0 for sale in SALES}
    pr_each = {sale: 0 for sale in SALES}
    pn_other = pv_other = pr_other = 0
    for s, qm in prev_short_quarter.items():
        if s in OLD24_SET: continue
        if not qm: continue
        if min(qm.keys()) != "2026Q2": continue
        owner = TUOKE_SHORT_TO_SALE.get(s)
        qc = qm.get("2026Q2", 0)
        is_valid = qc > 1000
        is_rising = is_valid and s in RISING_SET
        if owner in SALES:
            pn_each[owner] += 1
            if is_valid: pv_each[owner] += 1
            if is_rising: pr_each[owner] += 1
        else:
            pn_other += 1
            if is_valid: pv_other += 1
            if is_rising: pr_other += 1
        pn_total += 1
        if is_valid: pv_total += 1
        if is_rising: pr_total += 1
    # 给 new_summary 里的 2026Q2 行回填日Δ
    # 用户 2026-06-05 拍板改规矩：日Δ如实显示负数（之前2026-05-21定的clamp到0已废弃）
    # 原因：客户改简称→老客剔除等真实减少场景，clamp会掩盖真实变动；改为如实相减。
    def _clamp_pos(v):
        return v  # 不再clamp，如实显示（含负数）
    for entry in new_summary:
        if entry.get("quarter") != "2026Q2": continue
        sale = entry.get("sale")
        if sale == "\u5408\u8ba1":
            entry["newDayDelta"] = _clamp_pos(entry["newCount"] - pn_total)
            entry["validDayDelta"] = _clamp_pos(entry["validCount"] - pv_total)
            entry["risingDayDelta"] = _clamp_pos(entry["risingCount"] - pr_total)
        elif sale == "\u5176\u4ed6":
            entry["newDayDelta"] = _clamp_pos(entry["newCount"] - pn_other)
            entry["validDayDelta"] = _clamp_pos(entry["validCount"] - pv_other)
            entry["risingDayDelta"] = _clamp_pos(entry["risingCount"] - pr_other)
        elif sale in SALES:
            entry["newDayDelta"] = _clamp_pos(entry["newCount"] - pn_each.get(sale, 0))
            entry["validDayDelta"] = _clamp_pos(entry["validCount"] - pv_each.get(sale, 0))
            entry["risingDayDelta"] = _clamp_pos(entry["risingCount"] - pr_each.get(sale, 0))
    print(f"\n  \u65e5\u0394 \u5df2\u6ce8\u5165 center_quarter_summary.js (\u5408\u8ba1\u884c + 7\u9500\u552e + \u5176\u4ed6) - \u8d70 T-2 full_quarter \u91cd\u7b97\u5206\u652f")
elif prev_kpi_snap_file.exists():
    # 兜底：直接从 T-2 自己的 snapshot center_quarter_summary.js 读 2026Q2 行的 newCount/validCount/risingCount
    # 2026-05-22 修复 bug：用户每日给的 CSV 没 T-2 副本，必须靠 snapshot 才能算日Δ
    import re as _re
    snap_text = prev_kpi_snap_file.read_text(encoding='utf-8')
    m = _re.search(r"window\.__CENTER_QUARTER_SUMMARY__\s*=\s*(\[.*?\]);", snap_text, _re.S)
    prev_summary = json.loads(m.group(1)) if m else []
    prev_by_sale = {}
    for r in prev_summary:
        if r.get("quarter") == "2026Q2":
            prev_by_sale[r.get("sale", "")] = r
    def _clamp_pos(v):
        return v  # 2026-06-05改：不再clamp，如实显示负数
    for entry in new_summary:
        if entry.get("quarter") != "2026Q2": continue
        sale = entry.get("sale")
        prev = prev_by_sale.get(sale)
        if not prev: continue
        entry["newDayDelta"] = _clamp_pos(int(entry["newCount"]) - int(prev.get("newCount", 0)))
        entry["validDayDelta"] = _clamp_pos(int(entry["validCount"]) - int(prev.get("validCount", 0)))
        entry["risingDayDelta"] = _clamp_pos(int(entry["risingCount"]) - int(prev.get("risingCount", 0)))
    print(f"\n  \u65e5\u0394 \u5df2\u6ce8\u5165 center_quarter_summary.js (\u5408\u8ba1\u884c + 7\u9500\u552e + \u5176\u4ed6) - \u8d70 snapshot \u5151\u5e95\u5206\u652f T-2={PREV_SNAP_DATE}")
else:
    print(f"\n  \u26a0\ufe0f \u65e0 T-2 \u57fa\u51c6 (full_quarter_{T_MINUS_2}.csv \u548c snapshot {T_MINUS_2}/ \u90fd\u4e0d\u5b58\u5728) - \u65e5\u0394 \u5168 0")

# ===== R5.1 (2026-06-25): 注入新口径字段 yestCost环比 + cumYestCost累计昨耗 + qda季度日均 =====
# 字段定义见 重构方案_v2_2026-06-25/销售作战中心_系统重构方案_v2.md R5.1
#  - prevYestCost: T-2 同档 Q2 yestCost
#  - yestCostRate: (yestCost - prevYestCost) / prevYestCost
#  - cumYestCost: 25Q3+25Q4+26Q1+26Q2 同档 yestCost 合计 (今日)
#  - prevCumYestCost: 同上 (T-2)
#  - cumYestCostRate: (cumYestCost - prevCumYestCost) / prevCumYestCost
#  - qda: quarterCost / Q2_PASSED_DAYS
import re as _re2
from datetime import date as _date_r5

# R5 算 Q2 已过天数/剩余天数 (以 T-1 数据日为准, 含两端)
_q2_start_r5 = _date_r5(2026, 4, 1)
_q2_end_r5 = _date_r5(2026, 6, 30)
_t1_parts = T_MINUS_1.split("-")
_t1_date_r5 = _date_r5(int(_t1_parts[0]), int(_t1_parts[1]), int(_t1_parts[2]))
Q2_PASSED_DAYS = (_t1_date_r5 - _q2_start_r5).days + 1  # 含两端
Q2_REMAIN_DAYS = (_q2_end_r5 - _t1_date_r5).days + 1    # 含数据日 -> 末日
print(f"  R5.1 Q2 PASSED_DAYS={Q2_PASSED_DAYS} REMAIN_DAYS={Q2_REMAIN_DAYS} (\u542b\u4e24\u7aef)")

# 读 T-2 snapshot, 用于算 prevYestCost 和 prevCumYestCost
_r5_prev_snap_file = Path(f"{PC_DATA}/snapshots/{PREV_SNAP_DATE}/center_quarter_summary.js")
_r5_prev_by_sale_q2 = {}
_r5_prev_by_sale_cum = {}
_CUM_QS = ["2025Q3", "2025Q4", "2026Q1", "2026Q2"]
if _r5_prev_snap_file.exists():
    _snap_text2 = _r5_prev_snap_file.read_text(encoding='utf-8')
    _m2 = _re2.search(r"window\.__CENTER_QUARTER_SUMMARY__\s*=\s*(\[.*?\]);", _snap_text2, _re2.S)
    if _m2:
        _prev_full = json.loads(_m2.group(1))
        for _r in _prev_full:
            _q = _r.get("quarter")
            _s = _r.get("sale", "")
            if _q == "2026Q2":
                _r5_prev_by_sale_q2[_s] = _r
            if _q in _CUM_QS:
                _r5_prev_by_sale_cum[_s] = _r5_prev_by_sale_cum.get(_s, 0) + _r.get("yestCost", 0)

# 算今日各档 cumYestCost (25Q3+25Q4+26Q1+26Q2 合计)
_r5_now_by_sale_cum = {}
for _r in new_summary:
    _q = _r.get("quarter")
    _s = _r.get("sale", "")
    if _q in _CUM_QS:
        _r5_now_by_sale_cum[_s] = _r5_now_by_sale_cum.get(_s, 0) + _r.get("yestCost", 0)

def _safe_rate(a, b):
    return ((a - b) / b) if b else 0.0

_r5_injected = 0
for entry in new_summary:
    if entry.get("quarter") != "2026Q2": continue
    sale = entry.get("sale", "")
    prev_row = _r5_prev_by_sale_q2.get(sale)
    prev_yest = int(prev_row.get("yestCost", 0)) if prev_row else 0
    entry["prevYestCost"] = prev_yest
    entry["yestCostRate"] = round(_safe_rate(entry.get("yestCost", 0), prev_yest), 4)
    now_cum = _r5_now_by_sale_cum.get(sale, 0)
    prev_cum = _r5_prev_by_sale_cum.get(sale, 0)
    entry["cumYestCost"] = int(now_cum)
    entry["prevCumYestCost"] = int(prev_cum)
    entry["cumYestCostRate"] = round(_safe_rate(now_cum, prev_cum), 4)
    qcost = entry.get("quarterCost", 0)
    entry["qda"] = int(qcost / Q2_PASSED_DAYS) if Q2_PASSED_DAYS > 0 else 0
    _r5_injected += 1
print(f"  R5.1 \u5df2\u6ce8\u5165 yestCost\u73af\u6bd4/cumYestCost\u7d2f\u8ba1/qda \u5b57\u6bb5\u5230 {_r5_injected} \u884c Q2 \u6570\u636e")

out_cqs = f"{PC_DATA}/center_quarter_summary.js"
with open(out_cqs, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated 2026-05-15 from full_quarter+daily AData CSV (T-1={T_MINUS_1}) */\n")
    f.write("window.__CENTER_QUARTER_SUMMARY__ = ")
    f.write(json.dumps(new_summary, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
print(f"\n✅ center_quarter_summary.js: {os.path.getsize(out_cqs)/1024:.1f} KB")

# ============================================================
# 10. center_daily_kpi.js - 服饰中心日耗大卡
# ============================================================
q2_total = compute_segment("2026Q2", None)
q2_total_prev_q = compute_segment("2026Q1", None)
# 使用 CSV 自带的环比（5/16 vs 5/15），不自己算
center_chg = DAILY_TOTAL_RATE

# T-2 累计快照（如果存在 full_quarter_{T_MINUS_2}.csv 就算日Δ）
prev_day_snap_file = adata_dir / f"full_quarter_{T_MINUS_2}.csv"
prev_day_kpi = None
if prev_day_snap_file.exists():
    prev_short_quarter = defaultdict(dict)
    with open(str(prev_day_snap_file), encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
            if row[0].strip() in ('0', '1', '2'):
                row = row[1:]
            if len(row) < 3: continue
            short = row[0].strip()
            if not short or short == '\u6574\u4f53': continue
            q = to_quarter_from_str(row[1])
            cost = safe_float(row[2])
            if q and cost > 0:
                prev_short_quarter[short][q] = prev_short_quarter[short].get(q, 0) + cost
    pn = pv = pr = 0
    for s, qm in prev_short_quarter.items():
        if s in OLD24_SET: continue
        if not qm: continue
        if min(qm.keys()) != "2026Q2": continue
        pn += 1
        qc = qm.get("2026Q2", 0)
        if qc > 1000:
            pv += 1
            if s in RISING_SET:
                pr += 1
    prev_day_kpi = {'newCount': pn, 'validCount': pv, 'risingCount': pr}
    print(f"\n  5/15 \u5feb\u7167: \u65b0\u5ba2 {pn} / \u6709\u6548 {pv} / \u65b0\u9510 {pr}")
    print(f"  5/16 vs 5/15 \u65e5\u0394: \u65b0\u5ba2 {q2_total['newCount']-pn:+d} / \u6709\u6548 {q2_total['validCount']-pv:+d} / \u65b0\u9510 {q2_total['risingCount']-pr:+d}")

# 2026-05-25 兜底：full_quarter T-2 不存在但有 snapshot 时，从 snapshot 的合计行读
# 2026-06-26 子青拍板: 新锐名单换池后, 强制走 snapshot 同口径前一日 → 真实日Δ (不再用今日 RISING_SET 反算 T-2)
if prev_kpi_snap_file.exists():
    import re as _re2
    _txt = prev_kpi_snap_file.read_text(encoding='utf-8')
    _m = _re2.search(r"window\.__CENTER_QUARTER_SUMMARY__\s*=\s*(\[.*?\]);", _txt, _re2.S)
    if _m:
        _ps = json.loads(_m.group(1))
        for _r in _ps:
            if _r.get('quarter') == '2026Q2' and _r.get('sale') == '\u5408\u8ba1':
                prev_day_kpi = {
                    'newCount': int(_r.get('newCount', 0)),
                    'validCount': int(_r.get('validCount', 0)),
                    'risingCount': int(_r.get('risingCount', 0)),
                }
                print(f"  📸 snapshot {PREV_SNAP_DATE} \u5151\u5e95: \u65b0\u5ba2 {prev_day_kpi['newCount']} / \u6709\u6548 {prev_day_kpi['validCount']} / \u65b0\u9510 {prev_day_kpi['risingCount']}")
                break

center_daily_kpi = {
    "dataDate": T_MINUS_1,
    # kanban_embed.html 读取的字段（大写Y）
    "centerYestCost": round(DAILY_TOTAL_YEST, 1),
    "centerPrevCost": round(DAILY_TOTAL_PREV, 1),
    "centerDayCostRate": round(center_chg, 4),
    # mobile.html 读取的字段（兼容旧版）
    "centerCost": round(DAILY_TOTAL_YEST, 1),
    "centerCostChg": round(center_chg, 4),
    "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
    "source": "https://adata.woa.com/bi/view/15729?s=E0fBs",
    # 全量KPI
    "q2NewCount": q2_total["newCount"],
    "q2ValidCount": q2_total["validCount"],
    "q2RisingCount": q2_total["risingCount"],
    "q2NewYestCost": round(q2_total["yestCost"], 1),
    "q2NewDayCostRate": round(q2_total["dayCostRate"], 4),
    "q1NewCount": q2_total_prev_q["newCount"],
    "q1ValidCount": q2_total_prev_q["validCount"],
    "q1RisingCount": q2_total_prev_q["risingCount"],
}
# 注入日Δ（5/16 vs 5/15）
if prev_day_kpi:
    center_daily_kpi["prevDayNewCount"] = prev_day_kpi['newCount']
    center_daily_kpi["prevDayValidCount"] = prev_day_kpi['validCount']
    center_daily_kpi["prevDayRisingCount"] = prev_day_kpi['risingCount']
    center_daily_kpi["q2NewDayDelta"] = q2_total['newCount'] - prev_day_kpi['newCount']
    center_daily_kpi["q2ValidDayDelta"] = q2_total['validCount'] - prev_day_kpi['validCount']
    center_daily_kpi["q2RisingDayDelta"] = q2_total['risingCount'] - prev_day_kpi['risingCount']

# ===== R5.1 (2026-06-25): 注入 Q2 进度元数据 (剩余天/已过天/末日) =====
center_daily_kpi["q2PassedDays"] = Q2_PASSED_DAYS
center_daily_kpi["q2RemainDays"] = Q2_REMAIN_DAYS
center_daily_kpi["q2EndDate"] = "2026-06-30"
center_daily_kpi["q2TargetCost"] = 10920          # 万,死规矩(120w/d × 91d)
center_daily_kpi["q2TargetRise"] = 207
# 本季新客昨耗环比 (vs T-2 snapshot 顶部 q2NewYestCost)
_r5_prev_q2_new_yest = 0
try:
    if _r5_prev_snap_file.exists():
        # T-2 center_daily_kpi.js 同目录
        _prev_dk_file = Path(f"{PC_DATA}/snapshots/{PREV_SNAP_DATE}/center_daily_kpi.js")
        if _prev_dk_file.exists():
            _dk_text = _prev_dk_file.read_text(encoding='utf-8')
            _dk_m = _re2.search(r"window\.__CENTER_DAILY_KPI__\s*=\s*(\{.+?\});", _dk_text, _re2.S)
            if _dk_m:
                _prev_dk = json.loads(_dk_m.group(1))
                _r5_prev_q2_new_yest = _prev_dk.get("q2NewYestCost", 0)
except Exception as _e:
    print(f"  R5.1 warn: 读 T-2 center_daily_kpi 失败 {_e}")
center_daily_kpi["prevQ2NewYestCost"] = round(_r5_prev_q2_new_yest, 1)
center_daily_kpi["q2NewYestCostRate"] = round(
    ((q2_total["yestCost"] - _r5_prev_q2_new_yest) / _r5_prev_q2_new_yest) if _r5_prev_q2_new_yest else 0.0,
    4
)
print(f"  R5.1 daily_kpi: q2PassedDays={Q2_PASSED_DAYS} q2RemainDays={Q2_REMAIN_DAYS} q2NewYestCostRate={center_daily_kpi['q2NewYestCostRate']*100:+.2f}%")
out_dk = f"{PC_DATA}/center_daily_kpi.js"
with open(out_dk, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated 2026-05-15 (T-1={T_MINUS_1}) */\n")
    f.write("window.__CENTER_DAILY_KPI__ = ")
    f.write(json.dumps(center_daily_kpi, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
print(f"✅ center_daily_kpi.js: {os.path.getsize(out_dk)/1024:.1f} KB")
print(f"   服饰中心昨日消耗: {DAILY_TOTAL_YEST/10000:.1f}万  日环比: {center_daily_kpi['centerDayCostRate']*100:+.1f}%")
print(f"   Q2合计: 新客 {q2_total['newCount']} / 有效 {q2_total['validCount']} / 新锐 {q2_total['risingCount']} / 昨耗 {q2_total['yestCost']/10000:.1f}万")
print(f"   Q1对比: 新客 {q2_total_prev_q['newCount']} / 有效 {q2_total_prev_q['validCount']} / 新锐 {q2_total_prev_q['risingCount']}")

# ============================================================
# 10b. current_rising_data.js - 当前季度新锐名单（必须与 q2RisingCount 一致）
# ============================================================
current_rising = []
for _short in sorted(ALL_SHORTS):
    if _short in OLD24_SET:
        continue
    if compute_first_quarter(_short) != "2026Q2":
        continue
    if quarter_cost(_short, "2026Q2") <= 1000:
        continue
    if _short in RISING_SET:
        current_rising.append(_short)
out_cr = f"{PC_DATA}/current_rising_data.js"
with open(out_cr, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated {T_MINUS_1} from {loaded_from or 'RISING_SET'} */\n")
    f.write("window.__CURRENT_RISING_SET__ = new Set(")
    f.write(json.dumps(current_rising, ensure_ascii=False, separators=(',', ':')))
    f.write(");\n")
print(f"✅ current_rising_data.js: {len(current_rising)} 条（q2RisingCount={q2_total['risingCount']}）")
if len(current_rising) != q2_total['risingCount']:
    raise SystemExit(f"current_rising_data mismatch: set={len(current_rising)} q2RisingCount={q2_total['risingCount']}")

# ============================================================
# 11. delivery_side_summary.js
# ============================================================
delivery_counts = {
    'ADQ': len(ADQ_SET - QYT_SET - OLD24_SET),
    '\u5168\u57df\u901a': len(QYT_SET - ADQ_SET - OLD24_SET),
    '\u53cc\u6295\u653e': len(ADQ_SET & QYT_SET - OLD24_SET),
}
out_ds = f"{PC_DATA}/delivery_side_summary.js"
with open(out_ds, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated 2026-05-15 */\n")
    f.write(f"window.__DELIVERY_SIDE_MAP__ = {json.dumps(DELIVERY_SIDE_MAP, ensure_ascii=False, separators=(',',':'))};\n")
    f.write(f"window.__DELIVERY_COUNTS__ = {json.dumps(delivery_counts, ensure_ascii=False, separators=(',',':'))};\n")
print(f"✅ delivery_side_summary.js: {os.path.getsize(out_ds)/1024:.1f} KB")

# ============================================================
# 12. top80_effective_metrics.js (Top客户关键指标)
# ============================================================
topkey_file = adata_dir / "topkey_4yCkqO_response.csv"
TOP80_METRICS = {}
if topkey_file.exists():
    with open(str(topkey_file), encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        # 2026-06-03 二次修正列映射：6.3 关键指标 CSV 实测【每指标3列】(值/环比变化量/环比变化率%)，
        #   共16列。上一版误按每指标4列读(col4/5/8)，导致 costRate 取成了 ROI(全正小数),
        #   看板首页 cleanup-card statusOfRow 读 top80.costRate 全判稳定(0告警) → 待办状态全错。
        #   与 TOPKEY_RATES 段(L933)的3列映射统一。
        # [0]客户简称 [1]日均消耗 [2]消耗环比变化量 [3]消耗环比变化率%
        # [4]下单ROI [5]ROI环比变化量 [6]ROI环比变化率%
        # [7]CTR% [8]CTR变化量 [9]CTR变化率%
        # [10]浅层cvr% [11]cvr变化量 [12]cvr变化率%
        # [13]创意ID数 [14]创意ID变化量 [15]创意ID变化率%
        for row in reader:
            # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
            if row[0].strip() in ('0', '1', '2'):
                row = row[1:]
            if len(row) < 2: continue
            short = row[0].strip()
            if not short or short == '\u6574\u4f53': continue
            TOP80_METRICS[short] = {
                'cost': safe_float(row[1]) if len(row) > 1 else 0,
                'costRate': clean_rate(safe_float(row[3]) / 100) if len(row) > 3 else None,
                'roi': safe_float(row[4]) if len(row) > 4 else 0,
                'roiRate': clean_rate(safe_float(row[6]) / 100) if len(row) > 6 else None,
                'ctr': safe_float(row[7]) if len(row) > 7 else 0,
                'ctrRate': clean_rate(safe_float(row[9]) / 100) if len(row) > 9 else None,
                'cvr': safe_float(row[10]) if len(row) > 10 else 0,
                'cvrRate': clean_rate(safe_float(row[12]) / 100) if len(row) > 12 else None,
                'creativeCount': safe_float(row[13]) if len(row) > 13 else 0,
                'creativeRate': clean_rate(safe_float(row[15]) / 100) if len(row) > 15 else None,
            }

# 融合 daily 数据到 topkey
for short, m in TOP80_METRICS.items():
    d = DAILY.get(short, {})
    m['yestCost'] = d.get('yest', 0)
    m['prevCost'] = d.get('prev', 0)
    m['dayCostRate'] = (m['yestCost'] - m['prevCost']) / m['prevCost'] if m['prevCost'] > 0 else 0
    m['deliverySide'] = DELIVERY_SIDE_MAP.get(short, '')
    m['isNew'] = (compute_first_quarter(short) == '2026Q2') and (short not in OLD24_SET)
    m['isValid'] = m['isNew'] and quarter_cost(short, '2026Q2') > 1000
    m['isRising'] = m['isValid'] and short in RISING_SET
    # 销售归属
    owner = TUOKE_SHORT_TO_SALE.get(short)
    m['sale'] = owner if owner in SALES else '\u5176\u4ed6'

# 生成Top80数组：必须先是「本季度有效新客」(isValid=True)，按季度累计消耗降序取 Top10
TOP80_LIST = []
for sale in SALES:
    # 只保留 isValid=True 的客户
    sale_items = [(short, m) for short, m in TOP80_METRICS.items() if m['sale'] == sale and m.get('isValid')]
    sale_items.sort(key=lambda x: quarter_cost(x[0], '2026Q2'), reverse=True)
    for rank, (short, m) in enumerate(sale_items[:10], 1):
        qc = quarter_cost(short, '2026Q2')
        TOP80_LIST.append({
            'sale': sale, 'rank': rank, 'name': short,
            'deliverySide': m.get('deliverySide', ''),
            'yestCostWan': round(m.get('cost', 0) / 10000, 2),  # 改用 topkey.cost 口径，与 CSV 保持一致
            'quarterCost': round(qc, 1),
            'quarterCostWan': round(qc / 10000, 2),
            'bidCost': round(m.get('cost', 0), 1),
            'costRate': round_or_none(m.get('costRate'), 4),
            'roi': round(m.get('roi', 2), 2),
            'roiRate': round_or_none(m.get('roiRate'), 4),
            'creativeIds': round(m.get('creativeCount', 0), 0),
            'creativeRate': round_or_none(m.get('creativeRate'), 4),
            'ctr': round(m.get('ctr', 0), 2),
            'ctrRate': round_or_none(m.get('ctrRate'), 4),
            'cvr': round(m.get('cvr', 0), 2),
            'cvrRate': round_or_none(m.get('cvrRate'), 4),
            'isNew': m['isNew'], 'isValid': m['isValid'], 'isRising': m['isRising'],
        })
# 其他档：只保留 isValid=True 的客户，按季度累计消耗降序 Top10
other_items = [(short, m) for short, m in TOP80_METRICS.items() if m['sale'] == '\u5176\u4ed6' and m.get('isValid')]
other_items.sort(key=lambda x: quarter_cost(x[0], '2026Q2'), reverse=True)
for rank, (short, m) in enumerate(other_items[:10], 1):
    qc = quarter_cost(short, '2026Q2')
    TOP80_LIST.append({
        'sale': '\u5176\u4ed6', 'rank': rank, 'name': short,
        'deliverySide': m.get('deliverySide', ''),
        'yestCostWan': round(m.get('cost', 0) / 10000, 2),  # 改用 topkey.cost 口径，与 CSV 保持一致
        'quarterCost': round(qc, 1),
        'quarterCostWan': round(qc / 10000, 2),
        'bidCost': round(m.get('cost', 0), 1),
        'costRate': round_or_none(m.get('costRate'), 4),
        'roi': round(m.get('roi', 2), 2),
        'roiRate': round_or_none(m.get('roiRate'), 4),
        'creativeIds': round(m.get('creativeCount', 0), 0),
        'creativeRate': round_or_none(m.get('creativeRate'), 4),
        'ctr': round(m.get('ctr', 0), 2),
        'ctrRate': round_or_none(m.get('ctrRate'), 4),
        'cvr': round(m.get('cvr', 0), 2),
        'cvrRate': round_or_none(m.get('cvrRate'), 4),
        'isNew': m['isNew'], 'isValid': m['isValid'], 'isRising': m['isRising'],
    })

out_t80 = f"{PC_DATA}/top80_effective_metrics.js"
with open(out_t80, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated {datetime.now().strftime('%Y-%m-%d')} (T-1={T_MINUS_1}) */\n")
    f.write("window.__TOP80_EFFECTIVE_METRICS__ = ")
    f.write(json.dumps(TOP80_LIST, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
print(f"✅ top80_effective_metrics.js: {os.path.getsize(out_t80)/1024:.1f} KB  ({len(TOP80_LIST)} 条数组)")

# ============================================================
# 13. 够一够名单 - 有效新客差一步（首投季度消耗 500-1000）
# ============================================================
GYG_LIST = []
for short in ALL_SHORTS:
    if short in OLD24_SET: continue
    if compute_first_quarter(short) != '2026Q2': continue
    qc = quarter_cost(short, '2026Q2')
    if 500 < qc <= 1000:
        owner = TUOKE_SHORT_TO_SALE.get(short)
        sale = owner if owner in SALES else '\u5176\u4ed6'
        d = DAILY.get(short, {})
        GYG_LIST.append({
            'short': short,
            'sale': sale,
            'quarterCost': qc,
            'gap': 1000 - qc,
            'yestCost': d.get('yest', 0),
            'deliverySide': DELIVERY_SIDE_MAP.get(short, ''),
        })
GYG_LIST.sort(key=lambda x: x['gap'])

out_gyg = f"{PC_DATA}/enough_candidates.js"
with open(out_gyg, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated 2026-05-15 (T-1={T_MINUS_1}) */\n")
    f.write("window.__ENOUGH_CANDIDATES__ = ")
    f.write(json.dumps(GYG_LIST, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
print(f"✅ enough_candidates.js: {len(GYG_LIST)} 够一够客户（500-1000元）")

# ============================================================
# 14. 小红点 Top有效新客状态
# 2026-05-17 用户拍板：本季度有效新客 ∩ 各销售季度累计消耗 Top10 = 最多 70 客户
# 状态阈值不变
# ============================================================
topkey_file2 = adata_dir / "topkey_4yCkqO_response.csv"
TOPKEY_RATES = {}
if topkey_file2.exists():
    with open(str(topkey_file2), encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
            if row[0].strip() in ('0', '1', '2'):
                row = row[1:]
            if len(row) < 7: continue
            short = row[0].strip()
            if not short or short == '\u6574\u4f53': continue
            # 2026-06-03 修正列映射：关键指标-6.3.csv 是【每指标3列】结构：
            #   col0=客户简称 col1=日均消耗 col2=消耗环比变化量 col3=消耗环比变化率%
            #   col4=下单ROI col5=ROI环比变化量 col6=ROI环比变化率%
            #   （旧脚本误按4列读 col4/col5/col8，导致状态全0不触发告警）
            TOPKEY_RATES[short] = {
                'cost': safe_float(row[1]),
                'costRatePct': clean_pct(safe_float(row[3])),
                'roi': safe_float(row[4]),
                'roiRatePct': clean_pct(safe_float(row[6])),
            }

TARGET_Q = "2026Q2"
def is_q_valid_new(short):
    if short in OLD24_SET: return False
    if compute_first_quarter(short) != TARGET_Q: return False
    return quarter_cost(short, TARGET_Q) > 1000

TOP_STATUS = {}
TOP_STATUS_LIST = []
# 7 销售各 Top10
for sale in SALES:
    candidates = []
    for short in ALL_SHORTS:
        if not is_q_valid_new(short): continue
        owner = TUOKE_SHORT_TO_SALE.get(short)
        if owner != sale: continue
        candidates.append((short, quarter_cost(short, TARGET_Q)))
    candidates.sort(key=lambda x: x[1], reverse=True)
    for rank, (short, qc) in enumerate(candidates[:10], 1):
        rates = TOPKEY_RATES.get(short, {})
        cost_rate = rates.get('costRatePct')
        roi_rate = rates.get('roiRatePct')
        # 2026-06-02 脏值已清成 None，分类时视作无效(不触发告警)
        cr = cost_rate if cost_rate is not None else 0
        rr = roi_rate if roi_rate is not None else 0
        if cr < -20 or rr < -30:
            status = '\u9700\u544a\u8b66'
        elif cr < -10 or rr < -15:
            status = '\u9700\u5173\u6ce8'
        else:
            status = '\u57fa\u672c\u7a33\u5b9a'
        item = {
            'sale': sale, 'rank': rank, 'short': short,
            'quarterCost': round(qc, 1),
            'cost': rates.get('cost', 0),
            'costRatePct': cost_rate,
            'roi': rates.get('roi', 0),
            'roiRatePct': roi_rate,
            'status': status,
        }
        TOP_STATUS[short] = item
        TOP_STATUS_LIST.append(item)
# 其他档 Top10（非 7 销售归属）
others_candidates = []
for short in ALL_SHORTS:
    if not is_q_valid_new(short): continue
    owner = TUOKE_SHORT_TO_SALE.get(short)
    if owner in SALES: continue
    others_candidates.append((short, quarter_cost(short, TARGET_Q)))
others_candidates.sort(key=lambda x: x[1], reverse=True)
for rank, (short, qc) in enumerate(others_candidates[:10], 1):
    rates = TOPKEY_RATES.get(short, {})
    cost_rate = rates.get('costRatePct')
    roi_rate = rates.get('roiRatePct')
    # 2026-06-02 脏值已清成 None，分类时视作无效(不触发告警)
    cr = cost_rate if cost_rate is not None else 0
    rr = roi_rate if roi_rate is not None else 0
    if cr < -20 or rr < -30:
        status = '\u9700\u544a\u8b66'
    elif cr < -10 or rr < -15:
        status = '\u9700\u5173\u6ce8'
    else:
        status = '\u57fa\u672c\u7a33\u5b9a'
    item = {
        'sale': '\u5176\u4ed6', 'rank': rank, 'short': short,
        'quarterCost': round(qc, 1),
        'cost': rates.get('cost', 0),
        'costRatePct': cost_rate,
        'roi': rates.get('roi', 0),
        'roiRatePct': roi_rate,
        'status': status,
    }
    TOP_STATUS[short] = item
    TOP_STATUS_LIST.append(item)

out_ts = f"{PC_DATA}/top_status_data.js"
with open(out_ts, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated 2026-05-17 (T-1={T_MINUS_1}) - 7\u9500\u552e x Top10 = max 70 */\n")
    f.write("window.__TOP_STATUS_DATA__ = ")
    f.write(json.dumps(TOP_STATUS, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
    f.write("window.__TOP_STATUS_LIST__ = ")
    f.write(json.dumps(TOP_STATUS_LIST, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
warn = sum(1 for v in TOP_STATUS_LIST if v['status'] == '\u9700\u544a\u8b66')
watch = sum(1 for v in TOP_STATUS_LIST if v['status'] == '\u9700\u5173\u6ce8')
stable = sum(1 for v in TOP_STATUS_LIST if v['status'] == '\u57fa\u672c\u7a33\u5b9a')
print(f"\u2705 top_status_data.js: {len(TOP_STATUS_LIST)} \u5ba2\u6237 (7\u9500\u552e\u00d7Top10)  \u9700\u544a\u8b66:{warn} / \u9700\u5173\u6ce8:{watch} / \u57fa\u672c\u7a33\u5b9a:{stable}")

# ============================================================
# 14b. 小红点正向「起量机会·沉睡觉醒」(2026-06-02 用户拍板完整口径)
# 把小红点从"防漏"(负向告警)扩展到"抓机会"(正向起量)。
# 用 14 天日均做基线 + 绝对增量判定——从根上规避 costRate 百分比脏值坑。
# 死规矩——三道串联过滤,缺一不可(用户原话):
#   ① 有效新客 且 当天消耗(昨日) > 3000 元
#        有效新客 = 本季度(2026Q2)新客 且 非24年老客 且 季度累计消耗 > 1000
#   ② 当天 > 前一天(增量 > 0,在涨)
#   ③ 当天 - 近14天日均 > 2000 元(前两周几乎在睡觉,昨天突然觉醒放量)
# 14天日均来自用户从 adata 导出的 近14天日耗_YYYYMMDD.csv(列: 客户简称, 日均消耗(元))
# 遍历【全部有效新客】而非 Top10——潜力股量还没积累到能进 Top10,正是要抓的机会。
# ============================================================
# 加载用户导出的近14天日均(基线)
AVG14 = {}
_avg14_cands = sorted(
    glob.glob(str(adata_dir / "\u8fd114\u5929\u65e5\u8017*.csv")) +
    glob.glob("/Users/duziqing/Downloads/\u8fd114\u5929\u65e5\u8017*.csv") +
    glob.glob(f"{PROJECT}/\u8fd114\u5929\u65e5\u8017*.csv"),
    key=os.path.getmtime, reverse=True)
if _avg14_cands:
    with open(_avg14_cands[0], encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)  # 跳表头
        for row in reader:
            if len(row) < 2: continue
            s = row[0].strip()
            if not s or s == '\u6574\u4f53': continue
            AVG14[s] = safe_float(row[1])
    print(f"\u2705 14\u5929\u65e5\u5747: \u8bfb\u53d6 {len(AVG14)} \u5ba2\u6237  \u6765\u6e90 {os.path.basename(_avg14_cands[0])}")
else:
    print("\u26a0\ufe0f \u672a\u627e\u5230 \u8fd114\u5929\u65e5\u8017_*.csv,\u8d77\u91cf\u6bb5\u65e0\u6cd5\u7b97\u89c9\u9192\u589e\u91cf")

SURGE_MIN_COST = 3000      # ① 当天日耗门槛(元)
SURGE_WAKE_DELTA = 2000    # ③ 当天 - 14天日均 增量门槛(元)
SURGE_STRONG_COST = 10000  # 强起量当天绝对量阈值(元)

SURGE_LIST = []
_surge_skip_noavg = 0      # 命中前两条但无14天数据被跳过的数量(供核验)
for short in ALL_SHORTS:
    if compute_first_quarter(short) != '2026Q2':   # 本季度新客
        continue
    if short in OLD24_SET:                          # 非24年老客
        continue
    qc = quarter_cost(short, '2026Q2')
    if qc <= 1000:                                  # ① 有效新客(季度累计>1000)
        continue
    d = DAILY.get(short, {})
    yest = d.get('yest', 0)
    prev = d.get('prev', 0)
    if yest <= SURGE_MIN_COST:                      # ① 当天 > 3000
        continue
    if yest <= prev:                                # ② 在涨(增量>0)
        continue
    avg14 = AVG14.get(short)
    if avg14 is None:                               # 无14天基线,不猜,跳过
        _surge_skip_noavg += 1
        continue
    wake_delta = yest - avg14                        # 觉醒增量
    if wake_delta <= SURGE_WAKE_DELTA:               # ③ 当天 - 14天日均 > 2000
        continue
    is_strong = (yest >= SURGE_STRONG_COST)
    level = '\u5f3a\u8d77\u91cf' if is_strong else '\u6f5c\u529b\u8d77\u91cf'
    owner = TUOKE_SHORT_TO_SALE.get(short)
    sale = owner if owner in SALES else '\u5176\u4ed6'
    rates = TOPKEY_RATES.get(short, {})  # 复用 topkey 里的 ROI 等(可能为空)
    SURGE_LIST.append({
        'sale': sale,
        'short': short,
        'yestCost': round(yest, 1),
        'prevCost': round(prev, 1),
        'avg14': round(avg14, 1),                  # 14天日均(基线)
        'wakeDelta': round(wake_delta, 1),          # 觉醒增量 = 当天 - 14天日均
        'level': level,                             # 强起量 / 潜力起量
        'quarterCost': round(qc, 1),
        'roi': rates.get('roi', 0),
        'deliverySide': DELIVERY_SIDE_MAP.get(short, ''),
    })

# 排序: 强起量优先,组内按觉醒增量降序(觉醒幅度大的最该抓)
SURGE_LIST.sort(key=lambda x: (0 if x['level'] == '\u5f3a\u8d77\u91cf' else 1, -x['wakeDelta']))

out_surge = f"{PC_DATA}/top_rising_data.js"
with open(out_surge, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated {datetime.now().strftime('%Y-%m-%d')} (T-1={T_MINUS_1}) - \u8d77\u91cf\u673a\u4f1a\u00b7\u6c89\u7761\u89c9\u9192 */\n")
    f.write("window.__TOP_RISING_DATA__ = ")
    f.write(json.dumps(SURGE_LIST, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
strong = sum(1 for v in SURGE_LIST if v['level'] == '\u5f3a\u8d77\u91cf')
potential = sum(1 for v in SURGE_LIST if v['level'] == '\u6f5c\u529b\u8d77\u91cf')
print(f"\u2705 top_rising_data.js: {len(SURGE_LIST)} \u8d77\u91cf\u5ba2\u6237  \u5f3a\u8d77\u91cf:{strong} / \u6f5c\u529b\u8d77\u91cf:{potential}  (\u547d\u4e2d\u524d\u4e24\u6761\u4f46\u65e014\u5929\u57fa\u7ebf\u8df3\u8fc7:{_surge_skip_noavg})")

# ============================================================
# 13. 回写 tuoke_real_records.js 的明细字段（拓客报备明细每日刷新）
#     回写 9 个字段：deliverySide / yestCost / prevCost / costChg /
#                   quarterCost / isNew / isValid / isRising / status
#     口径与 compute_segment 保持一致：
#       - 新客 = first_quarter == 2026Q2 且 short ∉ OLD24_SET
#       - 有效 = 新客 且 quarterCost > 1000
#       - 新锐 = 有效 且 short ∈ RISING_SET
# ============================================================
TARGET_Q = "2026Q2"
records_path = f"{PC_DATA}/tuoke_real_records.js"
# 2026-06-03 兜底修复：当 REBUILD_OUTPUT_DIR 指向新空目录(如 out_clean)时，目标里没有
# tuoke 占位文件 → 下面 os.path.exists 判 False → 整段回填被静默跳过（报备明细不更新）。
# 修复：占位不存在则从真源 data 目录拷一份占位，保证回填始终触发。
if records and not os.path.exists(records_path):
    _src_tuoke = f"{PROJECT}/github_pages_adq_publish/sales-center/data/tuoke_real_records.js"
    if os.path.exists(_src_tuoke):
        import shutil as _sh
        os.makedirs(PC_DATA, exist_ok=True)
        _sh.copy(_src_tuoke, records_path)
        print(f"   ⚠️ {PC_DATA} 无 tuoke 占位 → 已从真源拷占位，确保回填触发")
if os.path.exists(records_path) and records:
    refresh_cnt = 0
    new_cnt = valid_cnt = rising_cnt = 0
    for r in records:
        short = (r.get("shortName") or r.get("name") or "").strip()
        if not short:
            continue
        # 投放端
        r["deliverySide"] = DELIVERY_SIDE_MAP.get(short, "")
        # 昨日 / 前日 / 日环比（直接读 daily CSV）
        d = DAILY.get(short, {})
        yest = round(d.get("yest", 0), 2)
        prev = round(d.get("prev", 0), 2)
        r["yestCost"] = yest
        r["prevCost"] = prev
        if prev > 0:
            r["costChg"] = f"{(yest - prev) / prev * 100:.1f}%"
        elif yest > 0:
            r["costChg"] = "新增"
        else:
            r["costChg"] = ""
        # 季度累计（本季度）
        qc = round(SHORT_QUARTER.get(short, {}).get(TARGET_Q, 0), 2)
        r["quarterCost"] = qc
        # 三层标
        fq = compute_first_quarter(short)
        is_new_base = (fq == TARGET_Q) and (short not in OLD24_SET)
        is_valid = bool(is_new_base and qc > 1000)
        is_rising = bool(is_valid and short in RISING_SET)
        r["isNew"] = is_new_base
        r["isValid"] = is_valid
        r["isRising"] = "\u662f" if is_rising else "\u5426"  # 是 / 否
        # R3.1.3 (2026-06-26): isLaoke 强制按 firstQuarter < 当前季度 计算, 保证与 firstQuarter 自洽
        # 之前 isLaoke 由 cloud_sync 写入, rebuild 不刷, 导致 30+ 条与 firstQuarter 自相矛盾
        if fq:
            r["isLaoke"] = bool(fq < TARGET_Q)
        # 2026-05-26 12:35 修复：firstQuarter 字段必须回写，否则前端「客户投放状态」
        # 只看这个字段会把 quarterCost=30万的本季新客误标为「新增未投放」
        # 写成 "2026/Q2" 格式（前端兼容 2026/Q2 和 2026Q2 两种写法）
        if fq:
            r["firstQuarter"] = f"{fq[:4]}/{fq[4:]}"  # "2026Q2" -> "2026/Q2"
        else:
            r["firstQuarter"] = ""
        # 状态文案：仅当本季度是真实首投 → 新客；老客名单 → 24老客；历史季度 → "xxxQx 新客"（标红）
        if short in OLD24_SET:
            r["status"] = "24\u8001\u5ba2"  # 24老客
        elif is_new_base:
            r["status"] = f"{TARGET_Q} \u65b0\u5ba2"  # 2026Q2 新客
        elif fq:
            r["status"] = f"{fq} \u65b0\u5ba2"  # 2025Q3 新客（历史季度新客，前端标红）
        else:
            r["status"] = "\u672a\u6295\u653e"  # 未投放
        refresh_cnt += 1
        if is_new_base: new_cnt += 1
        if is_valid: valid_cnt += 1
        if is_rising: rising_cnt += 1
    # 写回
    with open(records_path, "w", encoding="utf-8") as f:
        f.write(f"/* Auto-refreshed {T_MINUS_1} from rebuild_kpi_csv */\n")
        f.write("window.__TUOKE_REAL_RECORDS__ = ")
        f.write(json.dumps(records, ensure_ascii=False, separators=(',', ':')))
        f.write(";\n")
    print(f"\u2705 tuoke_real_records.js \u56de\u5199: {refresh_cnt} \u6761\uff08\u65b0{new_cnt}/\u6709\u6548{valid_cnt}/\u65b0\u9510{rising_cnt}\uff09")

# ============================================================
# 14. 每日重写 first_quarter.js（首投季度字典，投放状态判定的底层数据）
#     从 SHORT_QUARTER（已聚合 full_quarter CSV）反写 {short: "YYYY/Qx"}
# ============================================================
fq_path = f"{PC_DATA}/first_quarter.js"
fq_map = {}
for short, qs in SHORT_QUARTER.items():
    if not qs: continue
    # 取最早季度
    earliest = min(qs.keys())
    # 转回 "YYYY/Qx" 格式（前端就是这格式）
    m = re.match(r'(\d{4})Q(\d)', earliest)
    if m:
        fq_map[short] = f"{m.group(1)}/Q{m.group(2)}"
with open(fq_path, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated {T_MINUS_1} by rebuild_kpi_csv */\n")
    f.write("window.__FIRST_QUARTER__ = ")
    f.write(json.dumps(fq_map, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")
print(f"\u2705 first_quarter.js \u91cd\u5199: {len(fq_map)} \u4e2a\u5ba2\u6237\u9996\u6295\u5b63\u5ea6")

# ============================================================
# 15. 每日重写 customer_link_data.js（直播 / 小店 / 直播+小店 三态）
#     源：link_4695Or_response.csv
#     标记规则：
#       - 单链路 → 标该链路，bold=true
#       - 双链路（直播+小店）→ 标 "直播+小店"，消耗占比高的 bold=true
# ============================================================
link_csv = adata_dir / "link_4695Or_response.csv"
link_data = {}
if link_csv.exists():
    raw = defaultdict(list)
    with open(str(link_csv), encoding='utf-8-sig') as f:
        rd = csv.reader(f)
        next(rd)
        for row in rd:
            # 2026-05-28 新格式：第0列是 __level (0/2)，列偏移+1
            if row[0].strip() in ('0', '1', '2'):
                row = row[1:]
            if len(row) < 3: continue
            short = row[0].strip()
            link = row[1].strip()
            cost = safe_float(row[2])
            if not short or short == '\u6574\u4f53' or not link:
                continue
            raw[short].append({"link": link, "cost": cost})
    for short, arr in raw.items():
        if len(arr) == 1:
            arr[0]["share"] = 1.0
            arr[0]["bold"] = True
            link_data[short] = arr
        else:
            total = sum(x["cost"] for x in arr) or 1
            # 排序消耗降序
            arr.sort(key=lambda x: -x["cost"])
            arr[0]["bold"] = True
            for it in arr:
                it["share"] = round(it["cost"] / total, 4)
            for it in arr[1:]:
                it["bold"] = False
            link_data[short] = arr
    link_path = f"{PC_DATA}/customer_link_data.js"
    with open(link_path, "w", encoding="utf-8") as f:
        f.write(f"/* Auto-generated {T_MINUS_1} (T-1={T_MINUS_1}) by rebuild_kpi_csv */\n")
        f.write("window.__CUSTOMER_LINK_DATA__ = ")
        f.write(json.dumps(link_data, ensure_ascii=False, separators=(',', ':')))
        f.write(";\n")
    only_live = sum(1 for v in link_data.values() if len(v) == 1 and v[0]["link"] == "\u76f4\u64ad")
    only_shop = sum(1 for v in link_data.values() if len(v) == 1 and v[0]["link"] == "\u5c0f\u5e97")
    both = sum(1 for v in link_data.values() if len(v) >= 2)
    print(f"\u2705 customer_link_data.js \u91cd\u5199: {len(link_data)} \u4e2a\u5ba2\u6237\uff08\u4ec5\u76f4\u64ad{only_live}/\u4ec5\u5c0f\u5e97{only_shop}/\u53cc{both}\uff09")
else:
    print(f"\u26a0\ufe0f link_4695Or_response.csv \u4e0d\u5b58\u5728\uff0c\u8df3\u8fc7\u94fe\u8def\u5b57\u5178\u91cd\u5199")

# ============================================================
# 15.5 涨跌红黑榜 v2（redblack_data.js）
#     口径：非老客（!isLaoke && !old24），按销售分组，每销售每tab Top5涨+Top5跌
#     两个维度tab：本季度新客 / 历史累计新客
#     指标：客户/销售/昨日消耗/前日消耗/环比绝对值/环比百分比/季累/投放端/链路
# ============================================================
def _parse_cost_chg_num(chg_str):
    if not chg_str or chg_str == "\u65b0\u589e" or chg_str == "":
        return None
    try:
        return float(str(chg_str).replace("%", "").replace("+", "").strip())
    except (ValueError, TypeError):
        return None

def _get_link_str(short):
    links = link_data.get(short, [])
    if not links: return ""
    return "+".join(sorted(set(x.get("link", "") for x in links if x.get("link"))))

redblack_candidates = []
for r in records:
    # 2026-06-03 修复：原来用 `isLaoke or old24` 过滤，但 isLaoke 字段名误导——
    #   它的真实含义是「首投季度 ≠ 2026Q2」(底表 customerType 标的"老客 2025/Q3"等)，
    #   并非真老客。用它过滤会把 25Q3~26Q1 首投的历史新客(3184个)全部误剔，
    #   导致历史累计新客榜≈本季度榜(同一批26Q2客户)。
    #   正解：只排真·24年老客 OLD24_SET(old24)，历史季度首投新客必须保留。
    if r.get("old24"):
        continue
    chg_num = _parse_cost_chg_num(r.get("costChg", ""))
    if chg_num is None:
        continue
    if not r.get("yestCost") or r["yestCost"] <= 0:
        continue
    short = r.get("shortName", r.get("name", ""))
    redblack_candidates.append({
        "short": short, "sale": r.get("sale", ""),
        "yestCost": r.get("yestCost", 0), "prevCost": r.get("prevCost", 0),
        "costChg": r.get("costChg", ""), "_chgNum": chg_num,
        "quarterCost": r.get("quarterCost", 0), "isNew": r.get("isNew", False),
        "deliverySide": r.get("deliverySide", ""), "linkStr": _get_link_str(short),
        "cat": r.get("cat", ""), "firstQuarter": r.get("firstQuarter", ""),
    })

def _build_redblack_by_sale(rows, top_n=5):
    by_sale = {}
    for r in rows:
        sale = r.get("sale") or "\u5176\u4ed6"
        by_sale.setdefault(sale, []).append(r)
    result = {}
    for sale, items in by_sale.items():
        ups = sorted([r for r in items if r["_chgNum"] > 0], key=lambda x: x["_chgNum"], reverse=True)[:top_n]
        downs = sorted([r for r in items if r["_chgNum"] < 0], key=lambda x: x["_chgNum"])[:top_n]
        sale_rows = []
        for rank, r in enumerate(ups, 1):
            chg_abs = round(abs(r["yestCost"] - r["prevCost"]), 1) if r["prevCost"] > 0 else 0
            sale_rows.append({"rank": rank, "dir": "\u6da8", "short": r["short"], "sale": r["sale"],
                "yestCost": r["yestCost"], "prevCost": r["prevCost"],
                "costChgAbs": chg_abs, "costChgPct": round(r["_chgNum"], 1),
                "quarterCost": r["quarterCost"], "deliverySide": r["deliverySide"], "linkStr": r["linkStr"],
                "cat": r["cat"]})
        for rank, r in enumerate(downs, 1):
            chg_abs = round(abs(r["yestCost"] - r["prevCost"]), 1) if r["prevCost"] > 0 else 0
            sale_rows.append({"rank": rank, "dir": "\u8dcc", "short": r["short"], "sale": r["sale"],
                "yestCost": r["yestCost"], "prevCost": r["prevCost"],
                "costChgAbs": chg_abs, "costChgPct": round(r["_chgNum"], 1),
                "quarterCost": r["quarterCost"], "deliverySide": r["deliverySide"], "linkStr": r["linkStr"],
                "cat": r["cat"]})
        result[sale] = sale_rows
    return result

redblack_q2 = [r for r in redblack_candidates if r.get("isNew")]
# historical：25Q3起算的累计新客榜（25Q3 <= 首投季度 <= 26Q2，含本季度，与q2有重叠正常）
# 2026-06-03：用户拍板「25Q3起算」。候选池已只排真·24年老客(old24)，
#   故历史季度(25Q3/25Q4/26Q1)首投的新客现已保留，两榜不再恒等。
#   闭区间过滤防脏 firstQuarter（空串/未来季度/异常格式）混入。
def _fq_in_hist_range(r):
    fq = (r.get("firstQuarter") or "").replace("/", "")
    return "2025Q3" <= fq <= "2026Q2"
redblack_historical = [r for r in redblack_candidates if _fq_in_hist_range(r)]
redblack_data = {"q2": _build_redblack_by_sale(redblack_q2), "historical": _build_redblack_by_sale(redblack_historical)}

redblack_path = f"{PC_DATA}/redblack_data.js"
with open(redblack_path, "w", encoding="utf-8") as f:
    f.write(f"/* Auto-generated {T_MINUS_1} (T-1={T_MINUS_1}) by rebuild_kpi_csv */\n")
    f.write("window.__REDBLACK_DATA__ = ")
    f.write(json.dumps(redblack_data, ensure_ascii=False, separators=(',', ':')))
    f.write(";\n")

q2s = len(redblack_data["q2"]); q2t = sum(len(v) for v in redblack_data["q2"].values())
hs = len(redblack_data["historical"]); ht = sum(len(v) for v in redblack_data["historical"].values())
print(f"\u2705 redblack_data.js v2: \u672c\u5b63{q2s}\u4e2a\u9500\u552e{q2t}\u6761, \u5386\u53f2{hs}\u4e2a\u9500\u552e{ht}\u6761")

print("\n========== \u5168\u90e8\u5b8c\u6210 ==========")
print(f"输出目录: {PC_DATA}/")
for fn in ["center_quarter_summary.js", "center_daily_kpi.js", "delivery_side_summary.js",
           "top80_effective_metrics.js", "enough_candidates.js", "top_status_data.js",
           "tuoke_real_records.js", "first_quarter.js", "customer_link_data.js",
           "redblack_data.js"]:
    path = f"{PC_DATA}/{fn}"
    if os.path.exists(path):
        print(f"  ✅ {fn}: {os.path.getsize(path)/1024:.0f} KB")
    else:
        print(f"  ❌ {fn}: MISSING")
